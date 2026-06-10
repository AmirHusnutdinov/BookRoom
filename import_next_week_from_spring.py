import argparse
import re
import sqlite3
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Tuple

from openpyxl import load_workbook


IMAGE_URL = "https://web.archive.org/web/20220316020411im_/https://mipt.ru/upload/medialibrary/89f/eng_base_inversion.png"
ROOM_DESCRIPTION = "Аудитории, забронированные МФТИ на время проведение пар"


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def parse_room_header(value: str) -> Optional[Dict[str, object]]:
    text = normalize_text(value)
    if not text:
        return None

    # Ищем что-то вроде "801 КПМ"
    m = re.search(r"(\d{3,4}\s*[А-Яа-яA-Za-z]*)", text)
    if not m:
        return None

    room_name = m.group(1).strip()
    capacity = 1

    # Ищем первую скобку с числом, например "(13 ПК+6 ноут)"
    cap_match = re.search(r"\((\d+)", text)
    if cap_match:
        try:
            capacity = int(cap_match.group(1))
        except ValueError:
            capacity = 1

    return {
        "name": room_name,
        "capacity": capacity,
        "full_text": text,
    }


def parse_time_range(value: str) -> Optional[Tuple[str, str]]:
    text = normalize_text(value)
    if not text:
        return None

    text = text.replace("–", "-").replace("—", "-")
    m = re.search(r"(\d{1,2})[.:](\d{2})\s*-\s*(\d{1,2})[.:](\d{2})", text)
    if not m:
        return None

    sh, sm, eh, em = m.groups()
    start_time = "{:02d}:{:02d}".format(int(sh), int(sm))
    end_time = "{:02d}:{:02d}".format(int(eh), int(em))
    return start_time, end_time


def get_admin_user(conn: sqlite3.Connection) -> sqlite3.Row:
    conn.row_factory = sqlite3.Row
    admin = conn.execute(
        "SELECT * FROM users WHERE is_admin = 1 ORDER BY id LIMIT 1"
    ).fetchone()
    if not admin:
        raise RuntimeError("В базе не найден пользователь с is_admin = 1")
    return admin


def get_or_create_room(
    conn: sqlite3.Connection,
    room_name: str,
    capacity: int,
    dry_run: bool
) -> Optional[int]:
    conn.row_factory = sqlite3.Row

    room = conn.execute(
        "SELECT id FROM rooms WHERE name = ? LIMIT 1",
        (room_name,)
    ).fetchone()

    if room:
        return room["id"]

    if dry_run:
        print("[DRY-RUN] Будет добавлена аудитория:", room_name)
        return None

    cur = conn.execute(
        """
        INSERT INTO rooms (name, description, capacity, floor, equipment, image_url, is_active)
        VALUES (?, ?, ?, ?, ?, ?, 1)
        """,
        (room_name, ROOM_DESCRIPTION, capacity, 1, "", IMAGE_URL)
    )
    conn.commit()
    return cur.lastrowid


def booking_exists(
    conn: sqlite3.Connection,
    room_id: int,
    user_id: int,
    booking_date: str,
    start_time: str,
    end_time: str,
    purpose: str
) -> bool:
    conn.row_factory = sqlite3.Row
    row = conn.execute(
        """
        SELECT id
        FROM bookings
        WHERE room_id = ?
          AND user_id = ?
          AND booking_date = ?
          AND start_time = ?
          AND end_time = ?
          AND IFNULL(purpose, '') = IFNULL(?, '')
        LIMIT 1
        """,
        (room_id, user_id, booking_date, start_time, end_time, purpose)
    ).fetchone()
    return row is not None


def check_time_conflict(
    conn: sqlite3.Connection,
    room_id: int,
    booking_date: str,
    start_time: str,
    end_time: str
) -> bool:
    conn.row_factory = sqlite3.Row
    row = conn.execute(
        """
        SELECT COUNT(*) AS cnt
        FROM bookings
        WHERE room_id = ?
          AND booking_date = ?
          AND start_time < ?
          AND end_time > ?
        """,
        (room_id, booking_date, end_time, start_time)
    ).fetchone()
    return row["cnt"] > 0


def insert_booking(
    conn: sqlite3.Connection,
    room_id: int,
    user_id: int,
    booking_date: str,
    start_time: str,
    end_time: str,
    purpose: str,
    dry_run: bool
) -> bool:
    if room_id is None:
        print(
            "[DRY-RUN] Будет создана бронь:",
            booking_date, start_time, "-", end_time, "| room_id=<new room> |", purpose[:80]
        )
        return True

    if booking_exists(conn, room_id, user_id, booking_date, start_time, end_time, purpose):
        print(
            "[SKIP] Уже есть такая же бронь:",
            booking_date, start_time, "-", end_time, "| room_id={}".format(room_id)
        )
        return False

    if check_time_conflict(conn, room_id, booking_date, start_time, end_time):
        print(
            "[SKIP] Конфликт по времени:",
            booking_date, start_time, "-", end_time, "| room_id={}".format(room_id)
        )
        return False

    if dry_run:
        print(
            "[DRY-RUN] Будет создана бронь:",
            booking_date, start_time, "-", end_time, "| room_id={}".format(room_id), "|", purpose[:80]
        )
        return True

    conn.execute(
        """
        INSERT INTO bookings (room_id, user_id, booking_date, start_time, end_time, purpose)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (room_id, user_id, booking_date, start_time, end_time, purpose)
    )
    conn.commit()
    return True


def get_next_week_dates(reference_date: datetime) -> List[datetime]:
    # Следующая неделя от пятницы 24.04.2026 -> понедельник 27.04.2026
    days_to_next_monday = 7 - reference_date.weekday()
    if days_to_next_monday <= 0:
        days_to_next_monday += 7

    next_monday = reference_date + timedelta(days=days_to_next_monday)
    return [next_monday + timedelta(days=i) for i in range(7)]


def find_room_columns(ws) -> List[Dict[str, object]]:
    room_columns = []
    header_row = None

    max_scan_rows = min(ws.max_row, 15)
    for row in range(1, max_scan_rows + 1):
        found = 0
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            parsed = parse_room_header(value)
            if parsed:
                found += 1
        if found >= 3:
            header_row = row
            break

    if header_row is None:
        raise RuntimeError("Не удалось найти строку с аудиториями")

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        parsed = parse_room_header(value)
        if parsed:
            parsed["column"] = col
            room_columns.append(parsed)

    if not room_columns:
        raise RuntimeError("Не найдены аудитории на листе")

    return room_columns


def find_time_rows(ws) -> List[Dict[str, object]]:
    time_rows = []

    for row in range(1, ws.max_row + 1):
        row_time = None
        for col in range(1, min(ws.max_column, 5) + 1):
            value = ws.cell(row=row, column=col).value
            parsed = parse_time_range(value)
            if parsed:
                row_time = parsed
                break

        if row_time:
            time_rows.append({
                "row": row,
                "start_time": row_time[0],
                "end_time": row_time[1],
            })

    if not time_rows:
        raise RuntimeError("Не найдены временные слоты в левой части листа")

    return time_rows


def extract_cell_text(cell_value) -> str:
    text = normalize_text(cell_value)
    return text


def collect_schedule_entries(ws) -> List[Dict[str, object]]:
    room_columns = find_room_columns(ws)
    time_rows = find_time_rows(ws)

    entries = []
    for tr in time_rows:
        row = tr["row"]
        for room in room_columns:
            col = room["column"]
            value = ws.cell(row=row, column=col).value
            text = extract_cell_text(value)
            if not text:
                continue

            entries.append({
                "room_name": room["name"],
                "capacity": int(room.get("capacity", 1)),
                "purpose": text,
                "start_time": tr["start_time"],
                "end_time": tr["end_time"],
            })

    return entries


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Путь к xlsx-файлу")
    parser.add_argument("--db", required=True, help="Путь к SQLite базе")
    parser.add_argument("--sheet", required=True, help="Имя листа")
    parser.add_argument("--reference-date", required=True, help="Опорная дата в формате YYYY-MM-DD")
    parser.add_argument("--dry-run", action="store_true", help="Только показать, что будет сделано")
    args = parser.parse_args()

    ref_date = datetime.strptime(args.reference_date, "%Y-%m-%d")
    week_dates = get_next_week_dates(ref_date)

    print("Следующая неделя:")
    for d in week_dates:
        print("-", d.strftime("%Y-%m-%d"), "| weekday =", d.weekday())

    wb = load_workbook(args.xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise RuntimeError("Лист '{}' не найден. Доступные: {}".format(args.sheet, ", ".join(wb.sheetnames)))

    ws = wb[args.sheet]
    entries = collect_schedule_entries(ws)

    if not entries:
        print("На листе не найдено непустых записей расписания.")
        return

    conn = sqlite3.connect(args.db)
    conn.row_factory = sqlite3.Row

    admin = get_admin_user(conn)
    admin_id = admin["id"]
    admin_username = admin["username"]

    print("Админ:", admin_username, "(id={})".format(admin_id))
    print("Найдено шаблонных записей на листе:", len(entries))

    created_rooms = set()
    booked_count = 0
    skipped_count = 0

    # Логика: берем шаблон слотов с листа и размножаем на каждый день следующей недели
    for target_date in week_dates:
        booking_date = target_date.strftime("%Y-%m-%d")

        for entry in entries:
            room_name = entry["room_name"]
            capacity = entry["capacity"]
            purpose = entry["purpose"]
            start_time = entry["start_time"]
            end_time = entry["end_time"]

            room_id = get_or_create_room(conn, room_name, capacity, args.dry_run)
            if room_name not in created_rooms:
                created_rooms.add(room_name)

            ok = insert_booking(
                conn=conn,
                room_id=room_id,
                user_id=admin_id,
                booking_date=booking_date,
                start_time=start_time,
                end_time=end_time,
                purpose=purpose,
                dry_run=args.dry_run
            )
            if ok:
                booked_count += 1
            else:
                skipped_count += 1

    if not args.dry_run:
        conn.commit()
    conn.close()

    print()
    print("Готово.")
    print("Уникальных аудиторий обработано:", len(created_rooms))
    print("Броней создано/запланировано:", booked_count)
    print("Пропущено:", skipped_count)


if __name__ == "__main__":
    main()